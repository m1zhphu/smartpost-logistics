from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from core.database import get_db
from core.idempotency import validate_idempotency, commit_idempotency
from core.security import get_current_user
import crud.accounting as crud_acc 
from pydantic import BaseModel
from typing import List
from core.permissions import PermissionChecker
import schemas.accounting as schema_acc
import models
from io import BytesIO 
import pandas as pd 
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import decimal
from decimal import Decimal

router = APIRouter(prefix="/api/accounting", tags=["Accounting & Settlement"])

# --- SCHEMA CHO REQUEST ---
class CashConfirmRequest(BaseModel):
    waybill_codes: List[str]
    note: str = "Xác nhận Shipper đã nộp tiền mặt"

# --- 1. API LẤY DANH SÁCH CHỐT CA ---
@router.get("/cash-confirmation")
def get_cash_confirmation_list(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Lấy danh sách Shipper kèm tên thật từ hàm CRUD mới có JOIN bảng Users"""
    hub_id = current_user.get('primary_hub_id')
    return crud_acc.get_shippers_for_cash_confirmation(db, hub_id)

# --- 2. API XÁC NHẬN NỘP TIỀN (CHỐT CA) ---
@router.post("/confirm-shipper-cash")
async def confirm_shipper_cash(
    data: CashConfirmRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Xử lý chốt ca thu tiền từ Shipper"""
    try:
        hub_id = current_user.get('primary_hub_id')
        user_id = current_user.get('user_id')
        
        # Thực hiện cập nhật trạng thái và ghi Ledger
        count = crud_acc.record_cash_collection(db, data.waybill_codes, hub_id, user_id, data.note)
        
        return {"status": "SUCCESS", "message": f"Đã chốt thành công {count} đơn hàng"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- 3. CÁC API ĐỐI SOÁT SHOP & EXCEL ---
@router.post("/create-shop-statement")
async def create_shop_statement(
    customer_id: int, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    bills = crud_acc.get_settled_bills(db, customer_id)
    if not bills:
        raise HTTPException(status_code=404, detail="Không có đơn hàng nào chờ đối soát.")
    stmt = crud_acc.create_cod_settlement(db, customer_id, current_user.get('user_id'))
    if not stmt:
        raise HTTPException(status_code=400, detail="Lỗi tạo bảng kê")
    db.commit()
    return {
        "statement_id": stmt.statement_id,
        "download_url": f"/api/accounting/statements/{stmt.statement_id}/download",
        "data": bills
    }

# --- 4. API ĐỐI SOÁT KẾ TOÁN (STATEMENT DEBT) ---

def verify_accounting_access(user: dict):
    role_id = user.get("role_id")
    actual_role_id = user.get("actual_role_id")
    if role_id not in [1, 5] and actual_role_id not in [1, 5]:
        raise HTTPException(status_code=403, detail="Chỉ Admin và Kế toán mới có quyền truy cập chức năng này.")

@router.post("/statements", response_model=schema_acc.StatementDebtResponse)
def create_statement(
    data: schema_acc.StatementDebtCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Tạo bảng kê mới (DRAFT)"""
    verify_accounting_access(current_user)
    statement_data = data.model_dump()
    statement_data["created_by"] = current_user.get("user_id")
    
    statement = crud_acc.create_statement(db, statement_data)
    db.commit()
    return statement

@router.get("/statements/{statement_id}", response_model=schema_acc.StatementDebtResponse)
def get_statement_detail(
    statement_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Lấy chi tiết bảng kê (Đã bao gồm tính toán Adjustment)."""
    verify_accounting_access(current_user)
    statement = crud_acc.get_statement(db, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="Không tìm thấy bảng kê.")
    return statement

@router.patch("/statements/{statement_id}/status", response_model=schema_acc.StatementDebtResponse)
def update_statement_status(
    statement_id: int,
    data: schema_acc.StatementStatusUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Cập nhật trạng thái bảng kê. Nếu là CONFIRMED thì khóa giá các vận đơn."""
    verify_accounting_access(current_user)
    
    statement = crud_acc.get_statement(db, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="Không tìm thấy bảng kê.")
        
    if statement.status == "CONFIRMED" and data.status != "CONFIRMED":
        raise HTTPException(status_code=400, detail="Bảng kê đã CHỐT, không thể lùi trạng thái!")
        
    updated_statement = crud_acc.update_statement_status(db, statement, data.status)
    db.commit()
    return updated_statement

@router.post("/statements/cod", response_model=schema_acc.StatementResponse)
def create_cod_statement(
    data: schema_acc.StatementDebtCreate, # Dùng tạm schema create vì cấu trúc waybill_ids giống nhau
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Tạo bảng kê tiền thu hộ COD mới (DRAFT) - Mục 15 Đặc tả"""
    verify_accounting_access(current_user)
    
    # Gán người tạo
    statement_data = data.model_dump()
    statement_data["created_by"] = current_user.get("user_id")
    
    statement = crud_acc.create_cod_statement(db, statement_data)
    db.commit()
    db.refresh(statement)
    return statement

@router.post("/override-price")
def override_waybill_price(
    data: schema_acc.WaybillPriceOverride,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Sửa giá Vận đơn thủ công (Chỉ dành cho Kế toán). Bắt buộc ghi Log."""
    verify_accounting_access(current_user)
    
    # 1. Kiểm tra Vận đơn có nằm trong Bảng kê đã chốt không
    is_confirmed = crud_acc.is_waybill_in_confirmed_statement(db, data.waybill_id)
    
    waybill = db.query(models.Waybills).filter(models.Waybills.waybill_id == data.waybill_id).first()
    if not waybill:
        raise HTTPException(status_code=404, detail="Không tìm thấy Vận đơn.")

    if is_confirmed:
        # Mục 18: Nếu đã chốt, tạo adjustment thay vì sửa trực tiếp
        detail = db.query(models.StatementDetails).filter(models.StatementDetails.waybill_id == data.waybill_id).first()
        if not detail:
             raise HTTPException(status_code=400, detail="Vận đơn thuộc bảng kê đã chốt nhưng không tìm thấy mapping!")
             
        old_total = waybill.total_amount_to_collect or Decimal('0')
        new_shipping = data.new_shipping_fee if data.new_shipping_fee is not None else (waybill.shipping_fee or Decimal('0'))
        new_extra = data.new_extra_fee if data.new_extra_fee is not None else (waybill.extra_services_fee or Decimal('0'))
        
        # Tính VAT 8% và làm tròn
        new_total = ((new_shipping + new_extra) * Decimal('1.08')).quantize(Decimal('1'), rounding=decimal.ROUND_HALF_UP)
        
        diff = new_total - old_total
        
        adj = crud_acc.create_statement_adjustment(
            db, detail.statement_id, detail.statement_type, data.waybill_id, diff, data.reason, current_user.get("user_id")
        )
        db.commit()
        return {
            "status": "ADJUSTED",
            "message": "Bảng kê đã chốt. Đã tạo phiếu điều chỉnh thay vì sửa trực tiếp (Mục 18).",
            "adjustment_id": adj.id,
            "diff_amount": diff,
            "reason": data.reason
        }
        
    # Nếu chưa chốt, cho phép sửa trực tiếp (DRAFT mode)
    old_total = waybill.total_amount_to_collect or Decimal('0')
    
    if data.new_shipping_fee is not None:
        waybill.shipping_fee = data.new_shipping_fee
    if data.new_extra_fee is not None:
        waybill.extra_services_fee = data.new_extra_fee
        
    # Tính lại tổng tiền cho Vận đơn (Dùng Decimal)
    total_service = (waybill.shipping_fee or Decimal('0')) + (waybill.extra_services_fee or Decimal('0'))
    waybill.vat_amount = (total_service * Decimal('0.08')).quantize(Decimal('1'), rounding=decimal.ROUND_HALF_UP)
    waybill.total_amount_to_collect = total_service + waybill.vat_amount
    
    new_total = waybill.total_amount_to_collect
    
    # Ghi log Audit
    crud_acc.create_price_override_log(
        db=db,
        waybill_id=data.waybill_id,
        user_id=current_user.get("user_id"),
        old_amount=old_total,
        new_amount=new_total,
        reason=data.reason
    )
    
    db.commit()
    return {"message": "Cập nhật giá thành công (DRAFT mode)", "new_total": new_total}

@router.get("/cod/{statement_id}/export")
def export_statement_excel(
    statement_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    try:
        # GỌI CRUD thay vì viết query trực tiếp ở đây
        raw_data = crud_acc.get_statement_details_for_export(db, statement_id)

        if not raw_data:
            raise HTTPException(status_code=404, detail="Bảng kê không có dữ liệu chi tiết")

        processed_data = [
            {
                "Mã Vận Đơn": item.waybill_code,
                "Số Tiền (VNĐ)": float(item.amount),
                "Loại Bút Toán": "Tiền COD" if item.entry_type == "CREDIT" else "Phí/Dịch vụ",
                "Thời Gian": item.timestamp.strftime("%d/%m/%Y %H:%M") if item.timestamp else ""
            } for item in raw_data
        ]

        df = pd.DataFrame(processed_data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Doi_Soat_COD')
            worksheet = writer.sheets['Doi_Soat_COD']

            # Định nghĩa các Style
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            center_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # Áp dụng Style cho Header
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border

            # Tự động giãn cột và kẻ khung dữ liệu
            for column_cells in worksheet.columns:
                max_len = 0
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                        cell.border = thin_border
                        if cell.row > 1:
                            cell.alignment = Alignment(horizontal="left")
                    except: pass
                worksheet.column_dimensions[column_cells[0].column_letter].width = max_len + 5

        output.seek(0)
        headers = {
            'Content-Disposition': f'attachment; filename="Bang_ke_COD_{statement_id}.xlsx"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/statements/{statement_id}/export")
def export_debt_statement_excel(
    statement_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    try:
        raw_data = crud_acc.get_debt_statement_details_for_export(db, statement_id)

        if not raw_data:
            raise HTTPException(status_code=404, detail="Bảng kê không có dữ liệu chi tiết")

        processed_data = [
            {
                "Mã Vận Đơn": item.waybill_code,
                "Khách Hàng": item.company_name if item.company_name else "N/A",
                "Dịch Vụ": item.service_type,
                "TL Thực Tế (kg)": float(item.actual_weight) if item.actual_weight else 0,
                "TL Quy Đổi (kg)": float(item.converted_weight) if item.converted_weight else 0,
                "Cước Chính": float(item.shipping_fee) if item.shipping_fee else 0,
                "Phụ Phí": float(item.extra_services_fee) if item.extra_services_fee else 0,
                "VAT": float(item.vat_amount) if item.vat_amount else 0,
                "Thành Tiền": float(item.total_amount_to_collect) if item.total_amount_to_collect else 0
            } for item in raw_data
        ]

        df = pd.DataFrame(processed_data)
        
        # Thêm dòng tổng cộng
        if not df.empty:
            sums = df.select_dtypes(include=['number']).sum()
            total_row = {col: "" for col in df.columns}
            total_row["Mã Vận Đơn"] = "TỔNG CỘNG"
            for col in sums.index:
                total_row[col] = sums[col]
            df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Bang_Ke_Cuoc')
            worksheet = writer.sheets['Bang_Ke_Cuoc']

            # Định nghĩa các Style
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            center_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            # Định dạng số VNĐ
            currency_format = '#,##0'

            # Áp dụng Style cho Header
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border

            # Tự động giãn cột và kẻ khung dữ liệu
            for column_cells in worksheet.columns:
                max_len = 0
                col_letter = column_cells[0].column_letter
                col_name = worksheet.cell(row=1, column=column_cells[0].column).value

                for cell in column_cells:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                        cell.border = thin_border
                        
                        # Định dạng số cho các cột tiền
                        if col_name in ["Cước Chính", "Phụ Phí", "VAT", "Thành Tiền"] and cell.row > 1:
                            cell.number_format = currency_format
                            cell.alignment = Alignment(horizontal="right")
                        elif cell.row > 1:
                            cell.alignment = Alignment(horizontal="left")
                            
                        # Format đậm cho dòng tổng cộng
                        if cell.row == len(df) + 1:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    except: pass
                worksheet.column_dimensions[col_letter].width = max_len + 5

        output.seek(0)
        headers = {
            'Content-Disposition': f'attachment; filename="Bang_Ke_Cuoc_{statement_id}.xlsx"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/statements/{statement_id}/export-csv")
def export_debt_statement_csv(
    statement_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Xuất bảng kê cước ra file CSV (Mục 17 Đặc tả)"""
    verify_accounting_access(current_user)
    
    # 1. Lấy dữ liệu
    data = crud_acc.get_debt_statement_details_for_export(db, statement_id)
    if not data:
        raise HTTPException(status_code=404, detail="Không có dữ liệu để xuất.")
        
    # 2. Chuyển sang DataFrame
    df = pd.DataFrame(data, columns=[
        'Mã vận đơn', 'Khách hàng', 'Dịch vụ', 'Cân nặng', 
        'Cân quy đổi', 'Cước chính', 'Phụ phí', 'VAT (8%)', 'Tổng tiền'
    ])
    
    # 3. Tạo CSV stream
    output = BytesIO()
    df.to_csv(output, index=False, encoding='utf-8-sig')
    output.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="BangKe_Cuoc_{statement_id}.csv"'
    }
    return StreamingResponse(output, headers=headers, media_type='text/csv')

@router.get("/statements/{statement_id}/adjustments")
def get_statement_adjustments(
    statement_id: int,
    statement_type: str = "DEBT",
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Lấy danh sách phiếu điều chỉnh của bảng kê (Mục 18)"""
    verify_accounting_access(current_user)
    return crud_acc.get_adjustments_for_statement(db, statement_id, statement_type)

@router.get("/statements")
def list_statements(
    customer_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Lấy danh sách các bảng kê cước (DEBT) và COD đã lập"""
    verify_accounting_access(current_user)
    
    query_debt = db.query(models.StatementDebt)
    query_cod = db.query(models.StatementCOD)
    
    if customer_id:
        query_debt = query_debt.filter(models.StatementDebt.customer_id == customer_id)
        query_cod = query_cod.filter(models.StatementCOD.customer_id == customer_id)
        
    debts = query_debt.order_by(models.StatementDebt.statement_id.desc()).all()
    cods = query_cod.order_by(models.StatementCOD.statement_id.desc()).all()
    
    results = []
    for d in debts:
        results.append({
            "statement_id": d.statement_id,
            "statement_code": d.statement_code,
            "customer_id": d.customer_id,
            "grand_total": float(d.grand_total or 0),
            "status": d.status,
            "type": "DEBT",
            "created_at": d.created_at.isoformat() if d.created_at else None,
            "created_by": d.created_by
        })
    for c in cods:
        results.append({
            "statement_id": c.statement_id,
            "statement_code": c.statement_code,
            "customer_id": c.customer_id,
            "grand_total": float(c.total_amount or 0),
            "status": c.status,
            "type": "COD",
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "created_by": c.created_by
        })
        
    results.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return results

@router.get("/statements/{statement_id}/waybills")
def get_statement_waybills(
    statement_id: int,
    statement_type: str = "DEBT",
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Lấy danh sách vận đơn chi tiết trong bảng kê cước (DEBT) hoặc COD"""
    verify_accounting_access(current_user)
    
    details = db.query(models.StatementDetails).filter(
        models.StatementDetails.statement_id == statement_id,
        models.StatementDetails.statement_type == statement_type
    ).all()
    
    waybill_ids = [d.waybill_id for d in details]
    if not waybill_ids:
        return []
        
    waybills = db.query(models.Waybills).filter(models.Waybills.waybill_id.in_(waybill_ids)).all()
    return waybills